#!/usr/bin/env python3
"""
Radar Tìm Khách Hàng Facebook - Ngành Bảng Hiệu Quảng Cáo
Chỉ quét TIN MỚI + Tự động gửi Telegram
"""

import os
import sys
import json
import time
import argparse
from datetime import datetime, timedelta
from typing import List, Dict

from config import config
from keywords import keyword_engine
from scraper import FacebookRadarScraper
from analyzer import LeadAnalyzer
from notifier import LeadNotifier
from dashboard import LeadDashboard
from urllib.parse import quote
from openpyxl import load_workbook

class FacebookRadar: 
    
    def load_settings(self):

        settings = {}

        try:
            wb = load_workbook("settings.xlsx")
            ws = wb.active

            for row in ws.iter_rows(min_row=1, values_only=True):

                key = row[0]
                value = row[1]

                if key:
                    settings[str(key)] = value

            print("✅ Đã tải settings.xlsx")

        except Exception as e:

            print("❌ Không đọc được settings.xlsx")
            print(e)

        return settings

    def load_locations_from_excel(self):
        locations = []

        try:
            wb = load_workbook("locations.xlsx")
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):

                status = row[1]      # Cột B
                location = row[2]    # Cột C

                if status == "ON" and location:
                    locations.append(str(location).strip())

            print(f"✅ Đã tải {len(locations)} khu vực")

        except Exception as e:
            print("❌ Không đọc được locations.xlsx")
            print(e)

        return locations
    
    def load_keywords_from_excel(self):
        keywords = []

        try:
            wb = load_workbook("keywords.xlsx")
            ws = wb.active

            # Bỏ qua hàng tiêu đề
            for row in ws.iter_rows(min_row=2, values_only=True):
                keyword = row[1]   # Cột B

                if keyword:
                    keywords.append(str(keyword).strip())

            print(f"✅ Đã tải {len(keywords)} từ khóa từ Excel")

        except Exception as e:
            print("❌ Không đọc được keywords.xlsx")
            print(e)

        print("DEBUG Excel =", keywords)
        return keywords    
    
    def __init__(self):
        self.scraper = FacebookRadarScraper(
            access_token=config.FB_ACCESS_TOKEN,
            api_version=config.FB_API_VERSION
        )
        self.analyzer = LeadAnalyzer()
        self.notifier = LeadNotifier()
        self.dashboard = LeadDashboard()
        self.all_leads: List[Dict] = []

        self._load_existing_leads()
        
        self.scan_keywords = self.load_keywords_from_excel()
        self.locations = self.load_locations_from_excel()
        self.settings = self.load_settings()               

    def _load_existing_leads(self):
        data_file = "data/all_leads.json"
        if os.path.exists(data_file):
            try:
                with open(data_file, 'r', encoding='utf-8') as f:
                    self.all_leads = json.load(f)
                print(f"📂 Đã tải {len(self.all_leads)} leads từ lần chạy trước")
            except:
                print(self.locations)
                self.all_leads = []
                
                print(self.settings)
                

    def _save_all_leads(self):
        os.makedirs("data", exist_ok=True)
        with open("data/all_leads.json", 'w', encoding='utf-8') as f:
            json.dump(self.all_leads, f, ensure_ascii=False, indent=2)

    def _is_duplicate(self, post_id: str) -> bool:
        return any(lead.get('post_id') == post_id for lead in self.all_leads)

    def scan_keyword(self, keyword: str, hours_back: int = 24) -> List[Dict]:
        print(f"\n🔍 Đang quét: '{keyword}' (tin trong {hours_back}h qua)")

        if config.FB_ACCESS_TOKEN:
            posts = self.scraper.search_public_pages(keyword, limit=25, hours_back=hours_back)
        else:
            return self.scraper.simulate_manual_research([keyword])

            

        analyzed = self.analyzer.analyze_batch(posts)
        new_leads = []

        for lead in analyzed:
            lead_dict = lead.to_dict()
            if not self._is_duplicate(lead_dict['post_id']):
                self.all_leads.append(lead_dict)
                new_leads.append(lead_dict)

                # ⭐ GỬI TELEGRAM CHO HOT & WARM
                if lead_dict['lead_type'] in ['hot', 'warm']:
                    print(f"   🔔 Phát hiện {lead_dict['lead_type'].upper()} lead!")
                    self.notifier.notify_telegram(lead_dict)

                if lead_dict['lead_type'] == 'hot':
                    self.notifier.notify_console(lead_dict)

        if new_leads:
            print(f"   ✅ {len(new_leads)} leads mới")
        return new_leads

    def run_full_scan(self, hours_back: int = 24):
        print("="*60)
        print("🚀 RADAR FACEBOOK - CHỈ QUÉT TIN MỚI")
        print("="*60)
        print(f"⏰ Bắt đầu: {datetime.now().strftime('%H:%M:%S %d/%m/%Y')}")
        print(f"📅 Tin từ: {(datetime.utcnow() - timedelta(hours=hours_back)).strftime('%H:%M %d/%m')}")
        print(f"🔑 API: {'✅' if config.FB_ACCESS_TOKEN else '❌'} | 📱 Telegram: ✅")
        print("-"*60)

        # ⭐ TEST TELEGRAM TRƯỚC KHI QUÉT
        print("\n📱 Test Telegram trước...")
        self.notifier.test_telegram()

        scan_keywords = self.scan_keywords

        total_new = 0
        manual_links = []

        for keyword in scan_keywords:

            for location in self.locations:

                search_text = f"{keyword} {location}"

                print(f"🔍 {search_text}")

                result = self.scan_keyword(
                    search_text,
                    hours_back=hours_back
                )

                if config.FB_ACCESS_TOKEN:
                    total_new += len(result)
                else:
                    manual_links.extend(result)

                time.sleep(2)
        
        self._save_all_leads()
        # Gửi danh sách link lên Telegram (chỉ khi không dùng API)
        # Gửi danh sách link lên Telegram (tự chia nhiều tin nếu quá dài)
        if not config.FB_ACCESS_TOKEN and manual_links:

            header = (
                "🔎 FACEBOOK RADAR\n"
                f"🕒 {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
                f"📌 Tổng từ khóa: {len(manual_links)}\n\n"
            )

            message = header

            for i, item in enumerate(manual_links, start=1):

                line = (
                    f"{i}. {item['keyword']}\n"
                    f"👉 {item['facebook_search_url']}\n\n"
                )

                # Nếu sắp vượt giới hạn Telegram thì gửi trước
                if len(message) + len(line) > 3500:
                    self.notifier.send_message(message)
                    print("📲 Đã gửi 1 phần Telegram")
                    message = "🔎 FACEBOOK RADAR (tiếp)\n\n"

                message += line

            # Gửi phần còn lại
            if message.strip():
                self.notifier.send_message(message)

            print(f"📲 Đã gửi {len(manual_links)} link lên Telegram")
            print("DEBUG manual_links =", len(manual_links))

            print(f"📲 Đã gửi {len(manual_links)} link lên Telegram")
            print("DEBUG manual_links =", len(manual_links))
            print(manual_links)

        self._save_all_leads()
        

        if total_new > 0:
            hot_leads = [l for l in self.all_leads if l.get('lead_type') == 'hot']
            self.notifier.save_to_json(hot_leads, "hot_leads_latest.json")
            self.notifier.send_telegram_summary(
                [l for l in self.all_leads[-total_new:]],
                datetime.now().strftime('%H:%M %d/%m')
            )

        print("\n" + "="*60)
        self.dashboard.display_summary(self.all_leads)
        self.dashboard.display_hot_leads(self.all_leads, limit=5)
        self.dashboard.export_csv(self.all_leads)

        print(f"\n✅ Hoàn tất! {total_new} leads mới.")
        print("="*60)

    def run_interactive(self):

        print("📋 CHẾ ĐỘ INTERACTIVE")
        self.run_full_scan(
            hours_back=self.settings["scan_hours"]
        )
        
    def auto_run(self, interval_minutes=30):
        print("=" * 60)
        print("🤖 FACEBOOK RADAR AUTO")
        print("=" * 60)
        print(f"⏰ Quét mỗi {interval_minutes} phút")
        print("Nhấn Ctrl+C để dừng")
        print("=" * 60)

        try:
            while True:
                print(f"\n🕒 Bắt đầu quét: {datetime.now().strftime('%H:%M:%S')}")
                self.run_full_scan(
                    hours_back=self.settings["scan_hours"]
                )

                print(f"\n😴 Nghỉ {interval_minutes} phút...")
                time.sleep(interval_minutes * 60)

        except KeyboardInterrupt:
            print("\n👋 Đã dừng chương trình.")            


def main():
    parser = argparse.ArgumentParser(description='Radar Facebook')
    parser.add_argument('--mode', choices=['scan', 'interactive', 'monitor', 'test'],
                       default='interactive',
                       help='Chế độ: scan | interactive | monitor | test')
    parser.add_argument('--hours', type=int, default=24, help='Số giờ quét lùi')
    parser.add_argument('--interval', type=int, default=30, help='Phút giữa các lần quét')

    args = parser.parse_args()

    radar = FacebookRadar()
    if args.mode == 'test':
        print("🧪 TEST GỬI TELEGRAM")
        radar.notifier.test_telegram()

    elif args.mode == 'scan':
        radar.run_full_scan(hours_back=args.hours)

    elif args.mode == 'monitor':
        radar.continuous_monitor(args.interval)

    elif args.mode == 'interactive':
        radar.auto_run(60)

    else:
        print("❌ Mode không hợp lệ")

if __name__ == "__main__":
    main()